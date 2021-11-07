import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook

print("""
        [1]bütün hisselerin değerlerini çekmek istiyorsan
        [2]Kriter belirlemek istiyorsan
""")
while True:
    ##secenek = input("SEÇİMİNİZ: ")
    ##if int(secenek) == 1:
        sirketler_listesi = [#56 54
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/sayfalar/sirket-karti.aspx?hisse=AKSA",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AKSEN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALCAR",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALGYO",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALKA",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARDYZ",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARSAN",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ATLAS",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AVGYO",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARENA",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BLCYT",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BNTAS",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BRSAN",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BRYAT",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BUCIM",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=CEMTS",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=CIMSA",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=CMBTN",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DESPC",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DEVA",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DGATE",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ECZYT",
                            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGEEN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGPRO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGGUB",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGSER",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ENJSA",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ERBOS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ETYAT",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EUKYO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=FMIZP",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=FRIGO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GEDIK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GEDZA",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GENTS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GLBMD",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GLRYH",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GOODY",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GSDHO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=HATEK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=HDFGS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=INDES",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=INVEO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ISDMR",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ISMEN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ISYAT",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=JANTS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KERVT",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KFEIN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KLGYO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KONTR",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KORDS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KRVGD",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KUTPO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=LINK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=LKMNH",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MEGAP",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MTRYO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=NATEN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=NIBAS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=NUHCM",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OLMIP",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ORGE",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OSTIM",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OYLUM",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OZKGY",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PARSN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PEKGY",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PENGD",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PETUN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PNSUT",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=POLHO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PRKAB",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PRKME",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=RTALB",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=RYGYO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SANKO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SARKY",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SASA",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TATGD",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TLMAN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TMPOL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TMSN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TUKAS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ULUUN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=USAK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VERTU",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VKGYO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=YAYLA"
        ]
        sirketler_listesi_2=["https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AKCNS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALARK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALKIM",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARCLK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ASELS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BIMAS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DOAS",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EKGYO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ENKAI",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EREGL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KOZAA",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KOZAL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=LOGO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MAVI",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MPARK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OTKAR",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TAVHL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TCELL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=THYAO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TKFEN",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TOASO",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TTKOM",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TTRAK",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SELEC",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SAHOL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SISE",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ULKER",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VESBE",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VESTL",
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=YATAS"
                             ]#84 82
        sirketler_listesi_3=[#71,69
                             "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AKGRT",
        "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ANHYT",
        "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ANSGR"]

        wb = Workbook()

        ws = wb.active

        ws.title = "pdfk"

        ws = wb.create_sheet("PDFK")
        for i in sirketler_listesi:
            page = requests.get(i)
            soup = BeautifulSoup(page.content, 'html.parser')
            list_for_excel = []
            all_td_tags = []
            h1_tags = []
            for m in soup.select('h1'):
                h1_tags.append(m.text)
                print(h1_tags)
                list_for_excel.append(f"{m}")

            for j in soup.select('td'):
                all_td_tags.append(j.text)

            list_for_excel.append(all_td_tags[56])
            list_for_excel.append(all_td_tags[54])

            print(all_td_tags[56], all_td_tags[54])

            ws.append(list_for_excel)

        for i in sirketler_listesi_2:
            page = requests.get(i)
            soup = BeautifulSoup(page.content, 'html.parser')
            list_for_excel = []
            all_td_tags = []
            h1_tags = []
            for m in soup.select('h1'):
                h1_tags.append(m.text)
                print(h1_tags)
                list_for_excel.append(f"{m}")

            for j in soup.select('td'):
                all_td_tags.append(j.text)

            list_for_excel.append(all_td_tags[84])
            list_for_excel.append(all_td_tags[82])

            print(all_td_tags[84], all_td_tags[82])

            ws.append(list_for_excel)

        for i in sirketler_listesi_3:
            page = requests.get(i)
            soup = BeautifulSoup(page.content, 'html.parser')
            list_for_excel = []
            all_td_tags = []
            h1_tags = []
            for m in soup.select('h1'):
                h1_tags.append(m.text)
                print(h1_tags)
                list_for_excel.append(f"{m}")

            for j in soup.select('td'):
                all_td_tags.append(j.text)


            list_for_excel.append(all_td_tags[71])
            list_for_excel.append(all_td_tags[69])


            print(all_td_tags[71], all_td_tags[69])

            ws.append(list_for_excel)


        wb.save("pdfk.xlsx")
        wb.close()

    elif int(secenek) == 2:
        sirketler_listesi = [  # 56 54
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/sayfalar/sirket-karti.aspx?hisse=AKSA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AKSEN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALCAR",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALGYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALKA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARDYZ",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARSAN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ATLAS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AVGYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARENA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BLCYT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BNTAS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BRSAN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BRYAT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BUCIM",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=CEMTS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=CIMSA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=CMBTN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DESPC",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DEVA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DGATE",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ECZYT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGEEN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGPRO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGGUB",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EGSER",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ENJSA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ERBOS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ETYAT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EUKYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=FMIZP",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=FRIGO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GEDIK",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GEDZA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GENTS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GLBMD",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GLRYH",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GOODY",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=GSDHO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=HATEK",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=HDFGS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=INDES",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=INVEO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ISDMR",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ISMEN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ISYAT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=JANTS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KERVT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KFEIN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KLGYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KONTR",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KORDS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KRVGD",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KUTPO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=LINK",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=LKMNH",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MEGAP",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MTRYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=NATEN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=NIBAS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=NUHCM",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OLMIP",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ORGE",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OSTIM",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OYLUM",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OZKGY",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PARSN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PEKGY",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PENGD",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PETUN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PNSUT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=POLHO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PRKAB",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=PRKME",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=RTALB",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=RYGYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SANKO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SARKY",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SASA",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TATGD",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TLMAN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TMPOL",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TMSN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TUKAS",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ULUUN",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=USAK",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VERTU",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VKGYO",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=YAYLA"
        ]
        sirketler_listesi_2 = ["https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AKCNS",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALARK",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ALKIM",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ARCLK",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ASELS",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=BIMAS",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=DOAS",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EKGYO",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ENKAI",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=EREGL",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KOZAA",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=KOZAL",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=LOGO",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MAVI",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=MPARK",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=OTKAR",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TAVHL",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TCELL",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=THYAO",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TKFEN",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TOASO",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TTKOM",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=TTRAK",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SELEC",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SAHOL",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=SISE",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ULKER",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VESBE",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=VESTL",
                               "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=YATAS"
                               ]  # 84 82
        sirketler_listesi_3 = [  # 71,69
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=AKGRT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ANHYT",
            "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=ANSGR"]

        wb = Workbook()

        ws = wb.active

        ws.title = "pdfk"

        ws = wb.create_sheet("PDFK")
        pddd = input(int())
        pddd2 = input(int())
        for i in sirketler_listesi:
            page = requests.get(i)
            soup = BeautifulSoup(page.content, 'html.parser')
            list_for_excel = []
            all_td_tags = []
            h1_tags = []


            for j in soup.select('td'):
                all_td_tags.append(j.text)

            if all_td_tags[56] <= pddd2 and all_td_tags[56] >= pddd:
                for m in soup.select('h1'):
                    h1_tags.append(m.text)
                    print(h1_tags)
                    list_for_excel.append(f"{m}")
                print(all_td_tags[56], all_td_tags[54])
                list_for_excel.append(all_td_tags[56])
                list_for_excel.append(all_td_tags[54])
            else:
                pass

            #print(all_td_tags[56], all_td_tags[54])

            ws.append(list_for_excel)

        for i in sirketler_listesi_2:
            page = requests.get(i)
            soup = BeautifulSoup(page.content, 'html.parser')
            list_for_excel = []
            all_td_tags = []
            h1_tags = []

            for j in soup.select('td'):
                all_td_tags.append(j.text)

            if all_td_tags[84] <= pddd2 and all_td_tags[82] >= pddd:
                for m in soup.select('h1'):
                    h1_tags.append(m.text)
                    print(h1_tags)
                    list_for_excel.append(f"{m}")
                list_for_excel.append(all_td_tags[84])
                list_for_excel.append(all_td_tags[82])
                print(all_td_tags[84], all_td_tags[82])
            else:
                pass

            #print(all_td_tags[84], all_td_tags[82])

            ws.append(list_for_excel)

        for i in sirketler_listesi_3:
            page = requests.get(i)
            soup = BeautifulSoup(page.content, 'html.parser')
            list_for_excel = []
            all_td_tags = []
            h1_tags = []

            for j in soup.select('td'):
                all_td_tags.append(j.text)

            if all_td_tags[71] <= pddd2 and all_td_tags[71] >= pddd:
                for m in soup.select('h1'):
                    h1_tags.append(m.text)
                    print(h1_tags)
                    list_for_excel.append(f"{m}")
                print(all_td_tags[71], all_td_tags[69])
                list_for_excel.append(all_td_tags[71])
                list_for_excel.append(all_td_tags[69])
            else:
                pass

            #print(all_td_tags[71], all_td_tags[69])

            ws.append(list_for_excel)

        wb.save("pdfk.xlsx")
        wb.close()


    else:
        print("GİRDİĞİNİZ DEĞER GEÇERSİZ")
        break